"""
Data loading and validation module.
Handles Excel, CSV, and text input formats.
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime


class DataLoader:
    """Loads and validates input data from multiple formats."""

    REQUIRED_ABUSE_COLUMNS = ["date"]
    REQUIRED_COURT_COLUMNS = ["date"]

    def __init__(self, input_path=None):
        self.input_path = input_path
        self.raw_data = {}
        self.abuse_data = None
        self.court_data = None
        self.officials_data = None
        self.financial_data = None
        self.communications_data = None

    def load_excel(self, file_path):
        """Load all sheets from an Excel workbook."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{j}"
                               for j, h in enumerate(row)]
                else:
                    data.append(row)

            if headers and data:
                df = pd.DataFrame(data, columns=headers)
                df = self._clean_dataframe(df)
                sheet_data[sheet_name.lower().replace(" ", "_")] = df

        self.raw_data.update(sheet_data)
        wb.close()
        return sheet_data

    def load_csv(self, file_path):
        """Load a CSV file."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        df = pd.read_csv(file_path)
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        df = self._clean_dataframe(df)

        name = os.path.splitext(os.path.basename(file_path))[0].lower()
        self.raw_data[name] = df
        return {name: df}

    def load_text(self, text_content, name="pasted_data"):
        """Parse pasted text data (tab or comma delimited)."""
        lines = text_content.strip().split("\n")
        if not lines:
            return {}

        # Detect delimiter
        if "\t" in lines[0]:
            delimiter = "\t"
        elif "," in lines[0]:
            delimiter = ","
        else:
            delimiter = ","

        from io import StringIO
        df = pd.read_csv(StringIO(text_content), delimiter=delimiter)
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        df = self._clean_dataframe(df)

        self.raw_data[name] = df
        return {name: df}

    def load_directory(self, dir_path):
        """Load all supported files from a directory."""
        if not os.path.isdir(dir_path):
            raise NotADirectoryError(f"Not a directory: {dir_path}")

        all_data = {}
        for filename in sorted(os.listdir(dir_path)):
            filepath = os.path.join(dir_path, filename)
            ext = os.path.splitext(filename)[1].lower()

            if ext in (".xlsx", ".xls"):
                all_data.update(self.load_excel(filepath))
            elif ext == ".csv":
                all_data.update(self.load_csv(filepath))
            elif ext == ".txt":
                with open(filepath, "r") as f:
                    name = os.path.splitext(filename)[0].lower()
                    all_data.update(self.load_text(f.read(), name=name))

        return all_data

    def auto_classify_sheets(self):
        """Attempt to auto-classify loaded sheets by content."""
        classifications = {}
        for name, df in self.raw_data.items():
            cols_lower = set(str(c).lower() for c in df.columns)
            name_lower = name.lower()

            if any(k in name_lower for k in ["abuse", "incident", "behavior", "separation"]):
                self.abuse_data = df
                classifications[name] = "abuse_incidents"
            elif any(k in name_lower for k in ["official", "judge", "attorney", "gal", "guardian"]):
                self.officials_data = df
                classifications[name] = "court_officials"
            elif any(k in name_lower for k in ["court", "case", "filing", "motion", "docket"]):
                self.court_data = df
                classifications[name] = "court_records"
            elif any(k in name_lower for k in ["financ", "money", "cost", "fee", "payment"]):
                self.financial_data = df
                classifications[name] = "financial_records"
            elif any(k in name_lower for k in ["commun", "message", "email", "text", "call"]):
                self.communications_data = df
                classifications[name] = "communications"
            elif "abuse" in cols_lower or "incident" in cols_lower or "behavior" in cols_lower:
                self.abuse_data = df
                classifications[name] = "abuse_incidents"
            elif "judge" in cols_lower or "attorney" in cols_lower or "filing" in cols_lower:
                self.court_data = df
                classifications[name] = "court_records"
            else:
                classifications[name] = "unclassified"

        return classifications

    def _clean_dataframe(self, df):
        """Clean and standardize a dataframe."""
        # Drop completely empty rows and columns
        df = df.dropna(how="all").dropna(axis=1, how="all")

        # Try to parse date columns (skip columns with numeric-like names)
        skip_keywords = ["proximity", "count", "number", "num", "score",
                         "amount", "duration", "days", "hours", "id"]
        for col in df.columns:
            col_lower = str(col).lower()
            if any(s in col_lower for s in skip_keywords):
                continue
            if any(d in col_lower for d in ["date", "time", "when", "filed"]):
                # Only convert if values look like dates, not numbers
                sample = df[col].dropna().head(5)
                if len(sample) > 0 and all(isinstance(v, (int, float)) for v in sample):
                    continue
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                except Exception:
                    pass

        return df

    def get_summary(self):
        """Return a summary of all loaded data."""
        summary = {}
        for name, df in self.raw_data.items():
            summary[name] = {
                "rows": len(df),
                "columns": list(df.columns),
                "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                "null_counts": df.isnull().sum().to_dict(),
                "sample": df.head(3).to_dict()
            }
        return summary

    def merge_datasets(self, key_column="date"):
        """Attempt to merge datasets on a common key."""
        dfs_with_key = []
        for name, df in self.raw_data.items():
            if key_column in df.columns:
                df_copy = df.copy()
                df_copy["_source_sheet"] = name
                dfs_with_key.append(df_copy)

        if len(dfs_with_key) > 1:
            merged = pd.concat(dfs_with_key, ignore_index=True, sort=False)
            merged = merged.sort_values(key_column, na_position="last")
            return merged
        elif dfs_with_key:
            return dfs_with_key[0]
        return None
